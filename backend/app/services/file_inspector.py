import csv
from pathlib import Path
from typing import Any, Dict, Iterator, List, Union

from backend.app.config import get_settings
from backend.app.models.schemas import FileInspection
from backend.app.utils.file_utils import resolve_input_path
from backend.app.utils.hashing import sha256_file


SUPPORTED_EXTENSIONS = {".csv", ".tsv", ".xlsx"}


def inspect_file(file_path: Union[str, Path], max_preview_rows: int = None) -> FileInspection:
    path = resolve_input_path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Input file not found: {path}")

    extension = path.suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"Unsupported file extension: {extension}")

    preview_rows = max_preview_rows
    if preview_rows is None:
        preview_rows = get_settings().max_preview_rows

    if extension in {".csv", ".tsv"}:
        columns, row_count, preview = _inspect_delimited(path, preview_rows)
        metadata: Dict[str, Any] = {"delimiter": "," if extension == ".csv" else "\t"}
    else:
        columns, row_count, preview, sheet_name = _inspect_xlsx(path, preview_rows)
        metadata = {"sheet_name": sheet_name}

    return FileInspection(
        file_path=str(path),
        file_name=path.name,
        extension=extension,
        sha256=sha256_file(path),
        columns=columns,
        row_count=row_count,
        preview=preview,
        metadata=metadata,
    )


def iter_table_rows(file_path: Union[str, Path]) -> Iterator[Dict[str, Any]]:
    path = resolve_input_path(file_path)
    extension = path.suffix.lower()
    if extension in {".csv", ".tsv"}:
        yield from _iter_delimited_rows(path, "," if extension == ".csv" else "\t")
        return
    if extension == ".xlsx":
        yield from _iter_xlsx_rows(path)
        return
    raise ValueError(f"Unsupported file extension: {extension}")


def read_table_records(file_path: Union[str, Path]) -> List[Dict[str, Any]]:
    return list(iter_table_rows(file_path))


def _inspect_delimited(path: Path, max_preview_rows: int) -> tuple[List[str], int, List[Dict[str, Any]]]:
    row_count = 0
    preview: List[Dict[str, Any]] = []

    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="," if path.suffix.lower() == ".csv" else "\t")
        columns = list(reader.fieldnames or [])
        for row in reader:
            row_count += 1
            if len(preview) < max_preview_rows:
                preview.append(dict(row))

    return columns, row_count, preview


def _iter_delimited_rows(path: Path, delimiter: str) -> Iterator[Dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=delimiter)
        for row in reader:
            yield dict(row)


def _inspect_xlsx(path: Path, max_preview_rows: int) -> tuple[List[str], int, List[Dict[str, Any]], str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to inspect xlsx files") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [_normalize_cell(value) for value in next(rows, [])]
    row_count = 0
    preview: List[Dict[str, Any]] = []

    for values in rows:
        if not any(value is not None for value in values):
            continue
        row_count += 1
        if len(preview) < max_preview_rows:
            preview.append({headers[index]: value for index, value in enumerate(values[: len(headers)])})

    workbook.close()
    return headers, row_count, preview, sheet.title


def _iter_xlsx_rows(path: Path) -> Iterator[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read xlsx files") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [_normalize_cell(value) for value in next(rows, [])]
    try:
        for values in rows:
            if not any(value is not None for value in values):
                continue
            yield {headers[index]: value for index, value in enumerate(values[: len(headers)])}
    finally:
        workbook.close()


def _normalize_cell(value: Any) -> str:
    return "" if value is None else str(value).strip()

